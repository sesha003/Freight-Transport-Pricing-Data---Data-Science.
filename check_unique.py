import pandas as pd
import openpyxl

# Use openpyxl directly to read just the needed columns - much faster for large files
wb = openpyxl.load_workbook(r"D:\HFU UNI Project\DS-Project HFU\Cleaned_Transport_Data2.xlsx", read_only=True)
ws = wb.active

# Get headers
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

with open(r"D:\HFU UNI Project\DS-Project HFU\unique_results.txt", "w", encoding="utf-8") as f:
    f.write(f"All columns: {headers}\n\n")
    
    # Find indices for our target columns
    target_cols = {}
    for i, h in enumerate(headers):
        if h and any(k in str(h).lower() for k in ["source", "state"]):
            target_cols[h] = i
    
    f.write(f"Target columns found: {list(target_cols.keys())}\n\n")
    
    # Collect unique values
    unique_vals = {col: set() for col in target_cols}
    row_count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1
        for col_name, col_idx in target_cols.items():
            if col_idx < len(row) and row[col_idx] is not None:
                unique_vals[col_name].add(str(row[col_idx]))
    
    f.write(f"Total data rows: {row_count}\n\n")
    
    for col_name, vals in unique_vals.items():
        f.write(f"=== {col_name} ===\n")
        f.write(f"Unique count: {len(vals)}\n")
        sorted_vals = sorted(vals)
        for v in sorted_vals:
            f.write(f"  - {v}\n")
        f.write("\n")

wb.close()
print("DONE - results written to unique_results.txt")
